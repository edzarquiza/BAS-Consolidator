import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
import io
import copy
from datetime import datetime

st.set_page_config(
    page_title="Dexterous · BAS Consolidator",
    page_icon="📋",
    layout="wide"
)

st.markdown("""
<style>
  div.stButton > button {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
    color: white !important; font-weight: 600 !important;
    border: none !important; border-radius: 8px !important;
    width: 100%; padding: 0.55rem 1.2rem !important; font-size: 0.95rem !important;
  }
  div.stDownloadButton > button {
    background: linear-gradient(135deg, #38a169 0%, #2f855a 100%) !important;
    color: white !important; font-weight: 600 !important;
    border: none !important; border-radius: 8px !important;
    width: 100%; padding: 0.55rem 1.2rem !important; font-size: 0.95rem !important;
  }
  #MainMenu, footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

st.markdown(
    """<div style="background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
        border-radius:12px;padding:24px 32px 18px;margin-bottom:12px;
        display:flex;align-items:center;gap:18px;">
      <span style="font-size:2.2rem;">&#128203;</span>
      <div>
        <div style="color:#fff;font-size:1.45rem;font-weight:700;margin-bottom:3px;">BAS Consolidator</div>
        <div style="color:#a0aec0;font-size:0.86rem;">dexterous &middot; Bulk upload Xero exports &rarr; single BAS workbook</div>
      </div>
    </div>""",
    unsafe_allow_html=True,
)

# ── Constants ──────────────────────────────────────────────────────────────────
MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
CURRENT_YEAR = datetime.now().year
YEARS = [str(y) for y in range(CURRENT_YEAR - 2, CURRENT_YEAR + 3)]

REPORT_KEYWORDS = {
    "activity_statement": ["activity_statement", "activity statement"],
    "balance_sheet":      ["balance_sheet", "balance sheet"],
    "profit_and_loss":    ["profit_and_loss", "profit and loss", "profit_&_loss"],
    "payroll_activity":   ["payroll_activity_summary", "payroll activity summary", "payroll_activity", "payroll activity"],
    "aged_receivables":   ["aged_receivables_detail", "aged receivables detail", "aged_receivables", "aged receivables"],
    "aged_payables":      ["aged_payables_detail", "aged payables detail", "aged_payables", "aged payables"],
}

REPORT_LABELS = {
    "activity_statement": ("Activity Statement",       "GST Summary + GST Detail + BAS Field", True),
    "balance_sheet":      ("Balance Sheet",             "BS",                                   True),
    "profit_and_loss":    ("Profit & Loss",             "PL",                                   True),
    "payroll_activity":   ("Payroll Activity Summary",  "PAYROLL",                              False),
    "aged_receivables":   ("Aged Receivables Detail",   "AR",                                   False),
    "aged_payables":      ("Aged Payables Detail",      "AP",                                   False),
}

# ── Embedded masterlist ────────────────────────────────────────────────────────
DEFAULT_MASTERLIST = [
    {"Client Name": "Livewire Markets Pty Ltd",                        "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": "Accrual Basis"},
    {"Client Name": "PELAN PTY. LIMITED",                              "Status": "Inactive", "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": ""},
    {"Client Name": "THE TRUSTEE FOR SHEBELING TRUST",                 "Status": "Inactive", "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": ""},
    {"Client Name": "Sydney Region Aboriginal Corporation",            "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "Trupanion Australia Pty Ltd",                     "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": "Accrual Basis"},
    {"Client Name": "ISH DENTAL PTY LTD",                             "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "DEXTEROUS GROUP PTY LIMITED",                     "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": ""},
    {"Client Name": "TSM (THE SERVICE MANAGER) PTY LTD",              "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": ""},
    {"Client Name": "ELLIS ECOMMERCE PTY LTD",                        "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": "Cash Basis"},
    {"Client Name": "TWO BY FOUR CAFE PTY LTD",                       "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Quarterly",  "Accounting Method": "Cash Basis"},
    {"Client Name": "MARSILL PTY LTD",                                "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Quarterly",  "Accounting Method": "Accrual Basis"},
    {"Client Name": "HARMONY BUILD PTY LTD",                          "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Quarterly",  "Accounting Method": "Cash Basis"},
    {"Client Name": "FINTECH EQUITY PTY LTD",                         "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "THE TRUSTEE FOR SALMON FAMILY TRUST NO 2",       "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Accrual Basis"},
    {"Client Name": "MILLER,JOLYNN",                                   "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "THE TRUSTEE FOR LYNDCOTE FAMILY TRUST",          "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Quarterly",  "Accounting Method": "Accrual Basis"},
    {"Client Name": "EMMA MULHOLLAND ON HOLIDAY PTY LTD",             "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Quarterly",  "Accounting Method": "Cash Basis"},
    {"Client Name": "DOUGHBOY PIZZA FRANCHISING PTY LTD",             "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "HUNT CIVIL PTY LTD",                             "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": "Cash Basis"},
    {"Client Name": "SUPERFOODS AUSTRALIA PTY LTD",                   "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "3forward",                                        "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": "Accrual Basis"},
    {"Client Name": "WJ Dental Pty Ltd",                               "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "Bellinger Asset Management Pty Limited",          "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "Hansa Capital Pty Limited",                       "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "LYNDCOTE PTY LIMITED",                           "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "Stropro Operations Pty Ltd",                      "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "THE TRUSTEE FOR CAPSPACE TRUST 2024-1",          "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Accrual Basis"},
    {"Client Name": "Bellinger Credit Pty Limited",                    "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "Capspace Funds Management Pty Ltd",               "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "CAPSPACE PTY LTD",                               "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "CATO Location Services (QLD) Pty Ltd",           "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "Cato Logistics Pty Ltc",                          "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "CESSNOCK ZAM PTY LTD",                          "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "COFFS SOUTH ZAM PTY LTD",                       "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "ELLE FOOTBALL ACADEMY PTY LTD",                  "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Monthly",    "Accounting Method": "Cash Basis"},
    {"Client Name": "ENGINEERING LAB NSW PTY LTD",                    "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "PINNACLE HOSPITALITY GROUP PTY LTD",             "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "RICHBOX PTY LTD",                                "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "Stropro Technologies Pty Ltd",                    "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "The Trustee for Capspace Private Debt Fund",      "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "No Payroll", "Accounting Method": "Cash Basis"},
    {"Client Name": "The Trustee for Strobel Ryan Family Trust",       "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "The Trustee for The Studio by Abbey Hair Trust",  "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "BD Plus Pty Ltd",                                 "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "COOGEE ZAM PTY LTD",                            "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "PINNACLE CJ'S NOWRA PTY LTD",                   "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "PINNACLE CJ'S WARRAWONG PTY LTD",               "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "PINNACLE OPORTO ULLADULLA PTY LTD",              "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "PINNACLE OPORTO WARRAWONG PTY LTD",             "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "PINNACLE OPORTO WODONGA PTY LTD",               "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Cash Basis"},
    {"Client Name": "Spectrum Retail Developments",                    "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "The Trustee for Spectrum REIT 23 Unit Trust",    "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "THE TRUSTEE FOR SPECTRUM REIT 40 TRUST",        "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "",            "Accounting Method": "Accrual Basis"},
    {"Client Name": "UPSTAGE WORLD PTY LTD",                         "Status": "Active",   "Frequency": "Quarterly", "PAYG Instalment": "Quarterly",  "Accounting Method": "Cash Basis"},
]

# ── Session state: masterlist ──────────────────────────────────────────────────
if "masterlist" not in st.session_state:
    st.session_state.masterlist = [row.copy() for row in DEFAULT_MASTERLIST]

# ── Helpers ────────────────────────────────────────────────────────────────────
def detect_report_type(filename: str) -> str | None:
    name = filename.lower().replace(" ", "_")
    for rtype, keywords in REPORT_KEYWORDS.items():
        for kw in keywords:
            if kw.replace(" ", "_") in name:
                return rtype
    return None


def copy_sheet_data(src_ws, dst_ws):
    from openpyxl.cell.cell import MergedCell
    for row in dst_ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None
    for merged in list(dst_ws.merged_cells.ranges):
        dst_ws.unmerge_cells(str(merged))
    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            if cell.has_style:
                try:
                    dst_cell.font          = copy.copy(cell.font)
                    dst_cell.fill          = copy.copy(cell.fill)
                    dst_cell.border        = copy.copy(cell.border)
                    dst_cell.alignment     = copy.copy(cell.alignment)
                    dst_cell.number_format = cell.number_format
                except Exception:
                    pass
    # Copy column widths from source, then auto-widen any that are too narrow
    for col_letter, dim in src_ws.column_dimensions.items():
        try: dst_ws.column_dimensions[col_letter].width = dim.width
        except Exception: pass
    for row_idx, dim in src_ws.row_dimensions.items():
        try: dst_ws.row_dimensions[row_idx].height = dim.height
        except Exception: pass
    for merged in list(src_ws.merged_cells.ranges):
        try: dst_ws.merge_cells(str(merged))
        except Exception: pass

    # Auto-fit: widen columns where content exceeds the copied width
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell as _MC
    col_widths = {}
    for row in dst_ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MC) or cell.value is None:
                continue
            col_letter = get_column_letter(cell.column)
            cell_len = len(str(cell.value)) + 2
            col_widths[col_letter] = max(col_widths.get(col_letter, 0), cell_len)
    for col_letter, needed in col_widths.items():
        current = dst_ws.column_dimensions[col_letter].width or 8
        if needed > current:
            dst_ws.column_dimensions[col_letter].width = min(needed, 60)


def get_dst(out_wb, preferred, fallbacks=None):
    for n in [preferred] + (fallbacks or []):
        for sn in out_wb.sheetnames:
            if sn.strip().lower() == n.strip().lower():
                return out_wb[sn]
    return out_wb.create_sheet(preferred)


def _build_queries_sheet(ws, client_name, display_period, accounting_method, payg, file_name_val):
    """Write the Queries sheet with exact formatting matching the template."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── Colours ────────────────────────────────────────────────────────────────
    DARK_TEAL  = "FF76A5AF"   # label background
    LIGHT_TEAL = "FFA2C4C9"   # value background
    RED        = "FFFF0000"   # Note label
    BLACK      = "FF000000"

    def label_font(): return Font(name="Calibri", size=10, bold=True)
    def value_font(): return Font(name="Calibri", size=10, bold=False, color=BLACK)
    def plain_font(): return Font(name="Calibri", size=10, bold=False)
    def red_font():   return Font(name="Calibri", size=10, bold=True, color=RED)

    def dark_fill():  return PatternFill("solid", fgColor=DARK_TEAL)
    def light_fill(): return PatternFill("solid", fgColor=LIGHT_TEAL)

    def thin():   return Side(border_style="thin")
    def medium(): return Side(border_style="medium")
    def none():   return Side(border_style=None)

    def label_border():
        return Border(left=thin(), right=thin(), top=thin(), bottom=thin())

    def value_border_top():
        # B1:C1 — thick bottom
        return Border(left=thin(), right=thin(), top=thin(), bottom=medium())

    def value_border_mid():
        # B2:C2, B3:C3 — thick top + bottom
        return Border(left=thin(), right=none(), top=medium(), bottom=medium())

    def value_border_e1():
        return Border(left=thin(), right=none(), top=thin(), bottom=medium())

    def value_border_e23():
        return Border(left=thin(), right=none(), top=thin(), bottom=thin())

    def note_border():
        return Border(left=thin(), right=thin(), top=thin(), bottom=thin())

    left_align  = Alignment(horizontal="left")
    vcenter     = Alignment(horizontal="left", vertical="center")
    vtop        = Alignment(vertical="top")

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 17.89
    ws.column_dimensions["B"].width = 36.0
    ws.column_dimensions["C"].width = 1.11
    ws.column_dimensions["D"].width = 13.44
    ws.column_dimensions["E"].width = 12.55
    ws.column_dimensions["F"].width = 14.44

    # ── Row heights ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 12.75
    ws.row_dimensions[2].height = 14.4
    ws.row_dimensions[3].height = 14.4
    ws.row_dimensions[4].height = 12.75
    ws.row_dimensions[5].height = 13.2
    for r in range(6, 12):
        ws.row_dimensions[r].height = 13.8

    # ── Merge cells ────────────────────────────────────────────────────────────
    ws.merge_cells("B1:C1")
    ws.merge_cells("B2:C2")
    ws.merge_cells("B3:C3")

    # ── Row 1: Client Name / Period ────────────────────────────────────────────
    a1 = ws["A1"]; a1.value = "Client Name"
    a1.font = label_font(); a1.fill = dark_fill()
    a1.border = label_border(); a1.alignment = left_align

    b1 = ws["B1"]; b1.value = client_name
    b1.font = value_font(); b1.fill = light_fill()
    b1.border = value_border_top(); b1.alignment = left_align
    # C1 gets right border + top/bottom of merged cell
    c1 = ws["C1"]
    c1.border = Border(right=thin(), top=thin(), bottom=medium())

    d1 = ws["D1"]; d1.value = "Period"
    d1.font = label_font(); d1.fill = dark_fill()
    d1.border = label_border()

    e1 = ws["E1"]; e1.value = display_period
    e1.font = plain_font(); e1.fill = light_fill()
    e1.border = value_border_e1(); e1.alignment = left_align

    # ── Row 2: Accounting Method / Completed by ────────────────────────────────
    a2 = ws["A2"]; a2.value = "Accounting Method"
    a2.font = label_font(); a2.fill = dark_fill()
    a2.border = label_border(); a2.alignment = left_align

    b2 = ws["B2"]; b2.value = accounting_method
    b2.font = value_font(); b2.fill = light_fill()
    b2.border = value_border_mid(); b2.alignment = vcenter
    c2 = ws["C2"]
    c2.border = Border(top=medium(), bottom=medium())

    d2 = ws["D2"]; d2.value = "Completed by: "
    d2.font = label_font(); d2.fill = dark_fill()
    d2.border = label_border()

    e2 = ws["E2"]; e2.value = ""
    e2.font = value_font(); e2.fill = light_fill()
    e2.border = value_border_e23(); e2.alignment = left_align

    # ── Row 3: PAYG / Reviewed by ─────────────────────────────────────────────
    a3 = ws["A3"]; a3.value = "PAYG"
    a3.font = label_font(); a3.fill = dark_fill()
    a3.border = label_border(); a3.alignment = left_align

    b3 = ws["B3"]; b3.value = payg
    b3.font = value_font(); b3.fill = light_fill()
    b3.border = value_border_mid(); b3.alignment = vcenter
    c3 = ws["C3"]
    c3.border = Border(top=medium(), bottom=medium())

    d3 = ws["D3"]; d3.value = "Reviewed by: "
    d3.font = label_font(); d3.fill = dark_fill()
    d3.border = label_border()

    e3 = ws["E3"]; e3.value = ""
    e3.font = value_font(); e3.fill = light_fill()
    e3.border = value_border_e23(); e3.alignment = left_align

    # ── Row 4: File Name (no formatting) ──────────────────────────────────────
    ws["A4"].value = "File Name"; ws["A4"].font = plain_font()
    ws["B4"].value = file_name_val; ws["B4"].font = plain_font()

    # ── Row 5: Note (red bold, bordered) ──────────────────────────────────────
    a5 = ws["A5"]; a5.value = "Note"
    a5.font = red_font(); a5.border = note_border(); a5.alignment = left_align
    b5 = ws["B5"]
    b5.border = note_border()

    # ── Rows 6-8: plain text ──────────────────────────────────────────────────
    ws["A6"].value = "Email sent for queries and confrimation"
    ws["A6"].font = plain_font(); ws["A6"].alignment = vtop

    ws["A7"].value = "Subject reference"
    ws["A7"].font = plain_font(); ws["A7"].alignment = vtop

    ws["A8"].value = f"=+B4"
    ws["A8"].font = plain_font(); ws["A8"].alignment = vtop


def build_output(client_name, month, year, accounting_method, payg, frequency, report_wbs):
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    sheets_needed = ["Queries", "GST Summary", "GST Detail", "BAS field", "BS", "PL", "PAYROLL"]
    if accounting_method == "Cash Basis":
        sheets_needed += ["AR", "AP"]
    for sn in sheets_needed:
        out_wb.create_sheet(sn)

    if "activity_statement" in report_wbs:
        wb = report_wbs["activity_statement"]
        for sn in wb.sheetnames:
            ws = wb[sn]
            h = (ws.cell(1, 1).value or "").lower()
            if "activity statement" in h:
                copy_sheet_data(ws, get_dst(out_wb, "GST Summary"))
            elif "transactions by tax rate" in h:
                copy_sheet_data(ws, get_dst(out_wb, "GST Detail"))
            elif "transactions by bas field" in h:
                copy_sheet_data(ws, get_dst(out_wb, "BAS field", ["BAS Field"]))

    if "balance_sheet" in report_wbs:
        copy_sheet_data(report_wbs["balance_sheet"].active, get_dst(out_wb, "BS", ["BS "]))
    if "profit_and_loss" in report_wbs:
        copy_sheet_data(report_wbs["profit_and_loss"].active, get_dst(out_wb, "PL", ["P&L", "P&L "]))
    if "payroll_activity" in report_wbs:
        copy_sheet_data(report_wbs["payroll_activity"].active, get_dst(out_wb, "PAYROLL", ["Payroll Activity Smry"]))
    if "aged_receivables" in report_wbs and accounting_method == "Cash Basis":
        copy_sheet_data(report_wbs["aged_receivables"].active, get_dst(out_wb, "AR"))
    if "aged_payables" in report_wbs and accounting_method == "Cash Basis":
        copy_sheet_data(report_wbs["aged_payables"].active, get_dst(out_wb, "AP"))

    freq_suffix    = " Qtr" if frequency == "Quarterly" else ""
    display_period = f"{month}{year}_BAS{freq_suffix}"
    file_name_val  = f"{month}{year}_BAS {client_name}"

    queries_ws = None
    for sn in out_wb.sheetnames:
        if sn.lower() == "queries":
            queries_ws = out_wb[sn]
            break
    if queries_ws is None:
        queries_ws = out_wb.create_sheet("Queries", 0)

    _build_queries_sheet(queries_ws, client_name, display_period, accounting_method, payg, file_name_val)

    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue()


def masterlist_to_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "BAS"
    cols = ["Client Name", "Status", "Frequency", "PAYG Instalment", "Accounting Method"]
    ws.append(["XeroClientName", "Status", "Frequency", "PAYGI", "GSTAccountingMethod"])
    for row in rows:
        ws.append([row.get(c, "") for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# LAYOUT
# ══════════════════════════════════════════════════════════════════════════════
left, right = st.columns([1.05, 1], gap="large")

# ─────────────────────────────────────────────────────────────────────────────
# LEFT — Consolidator
# ─────────────────────────────────────────────────────────────────────────────
with left:
    st.subheader("① Client Details")

    col1, col2 = st.columns([2, 1])
    with col1:
        client_name_input = st.text_input("Client Name", placeholder="e.g. TSM THE SERVICE MANAGER PTY LTD")
    with col2:
        frequency = st.selectbox("Frequency", ["Quarterly", "Monthly"])

    col3, col4, col5 = st.columns(3)
    with col3:
        month_sel = st.selectbox("Month", MONTHS, index=MONTHS.index("DEC"))
    with col4:
        year_sel = st.selectbox("Year", YEARS, index=YEARS.index(str(CURRENT_YEAR)))
    with col5:
        accounting_method = st.selectbox("Accounting Method", ["Cash Basis", "Accrual Basis"])

    payg = st.selectbox("PAYG Instalment", ["Monthly", "Quarterly", "No Payroll"])

    # Auto-match from masterlist
    if client_name_input.strip():
        search = client_name_input.strip().upper()
        match = next(
            (r for r in st.session_state.masterlist if search in r["Client Name"].upper() or r["Client Name"].upper() in search),
            None
        )
        if match:
            parts = []
            if match["Frequency"]:        parts.append(f"Frequency: **{match['Frequency']}**")
            if match["PAYG Instalment"]:  parts.append(f"PAYG: **{match['PAYG Instalment']}**")
            if match["Accounting Method"]:parts.append(f"Method: **{match['Accounting Method']}**")
            if parts:
                st.success("✅ Found in masterlist — " + " · ".join(parts))

        freq_suffix  = " Qtr" if frequency == "Quarterly" else ""
        yr_short     = str(year_sel)[-2:]
        preview_name = f"{month_sel}{yr_short}_BAS{freq_suffix} {client_name_input.strip().upper()}.xlsx"
        st.info(f"📁 **{preview_name}**")

    st.divider()

    # ── Bulk Upload ───────────────────────────────────────────────────────────
    st.subheader("② Upload Xero Exports")
    st.caption("Drop all files at once — auto-detected from filename.")

    uploaded_files = st.file_uploader(
        "Select all Xero export files",
        type=["xlsx", "xlsm", "xls"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    detected   = {}
    undetected = []

    if uploaded_files:
        for f in uploaded_files:
            rtype = detect_report_type(f.name)
            if rtype:
                detected[rtype] = f
            else:
                undetected.append(f.name)

        st.markdown("**Detected files:**")
        for rtype, f in detected.items():
            label, sheet_dest, _ = REPORT_LABELS[rtype]
            st.markdown(f"- ✅ **{label}** → `{sheet_dest}`")
        if undetected:
            st.warning("⚠️ Unrecognised files: " + ", ".join(f"`{n}`" for n in undetected))

    st.divider()

    # ── Generate ──────────────────────────────────────────────────────────────
    st.subheader("③ Generate")

    ok_client = bool(client_name_input.strip())
    ok_files  = all(rt in detected for rt in ["activity_statement", "balance_sheet", "profit_and_loss"])
    all_ok    = ok_client and ok_files

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown("✅ Client"        if ok_client                          else "❌ Client name")
    c2.markdown("✅ Activity Stmt" if "activity_statement" in detected   else "❌ Activity Stmt")
    c3.markdown("✅ Balance Sheet" if "balance_sheet"       in detected  else "❌ Balance Sheet")
    c4.markdown("✅ P&L"          if "profit_and_loss"     in detected  else "❌ P&L")

    st.write("")
    if st.button("⚡  Generate BAS Workbook"):
        if not all_ok:
            st.error("Fill in client name and upload the 3 required files first.")
        else:
            with st.spinner("Consolidating…"):
                try:
                    report_wbs = {rtype: load_workbook(io.BytesIO(uf.read())) for rtype, uf in detected.items()}
                    output_bytes = build_output(
                        client_name=client_name_input.strip().upper(),
                        month=month_sel, year=year_sel,
                        accounting_method=accounting_method,
                        payg=payg, frequency=frequency,
                        report_wbs=report_wbs,
                    )
                    freq_suffix   = " Qtr" if frequency == "Quarterly" else ""
                    yr_short      = str(year_sel)[-2:]
                    download_name = f"{month_sel}{yr_short}_BAS{freq_suffix} {client_name_input.strip().upper()}.xlsx"
                    st.session_state["output_bytes"]  = output_bytes
                    st.session_state["download_name"] = download_name
                    st.success(f"✅ **{download_name}** ready!")
                except Exception as e:
                    import traceback
                    st.error(f"❌ {e}")
                    st.code(traceback.format_exc())

    if "output_bytes" in st.session_state:
        st.download_button(
            label="⬇  Download BAS Workbook",
            data=st.session_state["output_bytes"],
            file_name=st.session_state["download_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ─────────────────────────────────────────────────────────────────────────────
# RIGHT — Masterlist
# ─────────────────────────────────────────────────────────────────────────────
with right:
    st.subheader("📋 Client Masterlist")

    # ── Client match card (shown when left panel has a name typed) ────────────
    if client_name_input.strip():
        search_match = client_name_input.strip().upper()
        match = next(
            (r for r in st.session_state.masterlist
             if search_match in r["Client Name"].upper() or r["Client Name"].upper() in search_match),
            None
        )
        if match:
            freq   = match.get("Frequency") or "—"
            payg   = match.get("PAYG Instalment") or "—"
            method = match.get("Accounting Method") or "—"
            status = match.get("Status") or "—"
            st.markdown(
                f"""<div style="background:#1a4731;border:1px solid #2f855a;border-radius:8px;
                    padding:12px 16px;margin-bottom:10px;">
                  <div style="color:#68d391;font-size:0.75rem;font-weight:700;
                      text-transform:uppercase;letter-spacing:0.06em;margin-bottom:6px;">
                    ✅ Matched Client
                  </div>
                  <div style="color:#f0fff4;font-size:0.95rem;font-weight:700;margin-bottom:8px;">
                    {match["Client Name"]}
                  </div>
                  <div style="display:flex;gap:20px;flex-wrap:wrap;">
                    <span style="color:#9ae6b4;font-size:0.8rem;">
                      <span style="color:#68d391;font-weight:600;">Frequency</span><br>{freq}
                    </span>
                    <span style="color:#9ae6b4;font-size:0.8rem;">
                      <span style="color:#68d391;font-weight:600;">PAYG</span><br>{payg}
                    </span>
                    <span style="color:#9ae6b4;font-size:0.8rem;">
                      <span style="color:#68d391;font-weight:600;">Method</span><br>{method}
                    </span>
                    <span style="color:#9ae6b4;font-size:0.8rem;">
                      <span style="color:#68d391;font-weight:600;">Status</span><br>{status}
                    </span>
                  </div>
                </div>""",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"""<div style="background:#2d1b1b;border:1px solid #c53030;border-radius:8px;
                    padding:10px 16px;margin-bottom:10px;">
                  <span style="color:#fc8181;font-size:0.85rem;">
                    ⚠️ <strong style="color:#feb2b2;">{client_name_input.strip().upper()}</strong>
                    not found in masterlist
                  </span>
                </div>""",
                unsafe_allow_html=True,
            )

    # ── Search filter ─────────────────────────────────────────────────────────
    search_q = st.text_input(
        "Search", placeholder="🔍  Filter by client name...",
        label_visibility="collapsed"
    )

    STATUS_OPTS = ["Active", "Inactive"]
    FREQ_OPTS   = ["Quarterly", "Monthly"]
    PAYG_OPTS   = ["Monthly", "Quarterly", "No Payroll", ""]
    METHOD_OPTS = ["Cash Basis", "Accrual Basis", ""]

    full_df = pd.DataFrame(st.session_state.masterlist)

    if search_q.strip():
        display_df = full_df[
            full_df["Client Name"].str.upper().str.contains(search_q.strip().upper(), na=False)
        ].reset_index(drop=True)
    else:
        display_df = full_df.copy()

    st.caption("Click any cell to edit · Scroll to bottom to add a new row · Changes save automatically.")

    edited_df = st.data_editor(
        display_df,
        use_container_width=True,
        hide_index=True,
        height=400,
        num_rows="dynamic",
        column_config={
            "Client Name": st.column_config.TextColumn(
                "Client Name", width="large", required=True,
            ),
            "Status": st.column_config.SelectboxColumn(
                "Status", options=STATUS_OPTS, width="small", required=True,
            ),
            "Frequency": st.column_config.SelectboxColumn(
                "Frequency", options=FREQ_OPTS, width="small", required=True,
            ),
            "PAYG Instalment": st.column_config.SelectboxColumn(
                "PAYG Instalment", options=PAYG_OPTS, width="small",
            ),
            "Accounting Method": st.column_config.SelectboxColumn(
                "Accounting Method", options=METHOD_OPTS, width="medium",
            ),
        },
        key="masterlist_editor",
    )

    # Persist edits back to session state
    if edited_df is not None:
        edited_records = edited_df.fillna("").to_dict("records")
        edited_records = [r for r in edited_records if str(r.get("Client Name", "")).strip()]
        if search_q.strip():
            q = search_q.strip().upper()
            unmatched = [r for r in st.session_state.masterlist if q not in r["Client Name"].upper()]
            st.session_state.masterlist = unmatched + edited_records
        else:
            st.session_state.masterlist = edited_records

    active_count = sum(1 for r in st.session_state.masterlist if r.get("Status") == "Active")
    st.caption(f"{len(st.session_state.masterlist)} clients total · {active_count} active")

    # ── Download updated tracker ──────────────────────────────────────────────
    st.divider()
    st.download_button(
        label="⬇  Download Updated Masterlist (.xlsx)",
        data=masterlist_to_xlsx(st.session_state.masterlist),
        file_name="BAS_Monthly_Automation_Tracker_Updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.divider()
st.caption("dexterous · BAS Consolidator")
